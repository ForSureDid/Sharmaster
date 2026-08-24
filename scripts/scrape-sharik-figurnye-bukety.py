#!/usr/bin/env python3
"""
Scrapes all products from the sharik.ru "Шары фигурные (букеты)" category into a
plain data-export xlsx (NOT the "Template For Sharik Ru.xlsx" format — this is a
one-off 9-column export, no image upload / DB linking).

URL: https://new.sharik.ru/tovary-dly-prazdnika-cat/folgirovannye-shary/shary-figurnye-bukety/

Two-phase, resumable:
  1. Crawl category listing pages -> collect {id, slug, code} for every product
     (checkpoint: listing.json)
  2. Fetch each product detail page -> extract fields from window.__INITIAL_STATE__
     (checkpoint: raw.jsonl, append-only, one JSON object per line)
Then builds the final xlsx from raw.jsonl.

Field notes (confirmed by inspecting window.__INITIAL_STATE__ on real product pages
before writing this — see conversation record for the 3 sample dumps):
  - Артикул       = product["code"] (sharik.ru's own SKU, e.g. "1207-7166" — this is
                     what's shown/used as the article on the page; distinct from
                     "Артикул производителя" which is the manufacturer's own code and
                     is NOT what this column captures).
  - Размеры/Габариты, Цвет фольга, Коллекция  = from product["properties"]
    (Размеры/Габариты's real property name has a literal backslash: "Размеры\\Габариты").
  - Торговая марка = from product["origin_properties"]. Absent (blank) for generic
    unbranded Chinese figures (the common "К ФИГУРА ..." / "Фигура Китай" group)
    — that's real, not a bug.
  - Вес брутто / Размер ("Вид упаковки" tab) = from product["measure_units_properties"]
    (rows keyed by name, e.g. "Вес брутто", "Размер") x product["measure_units_names"]
    (columns — the packaging units for this product, e.g. ["Штука","Блок","Коробка"]).
    There is no separate "packing/upakovka" JSON block; this measure_units_* pair IS
    the "Вид упаковки" tab data, confirmed by diffing against the raw pre-hydration
    HTML (the tab is 100% client-rendered from this state, nothing to cross-check in
    static markup).
    Selection rule (CORRECTED 2026-08-24 — per-piece value wanted, not packaged):
    priority order Штука > Блок > Коробка. The task explicitly wants the weight/size
    of ONE PIECE. "Блок" and "Коробка" are packaging/carton weights covering many
    pieces, which is wrong for this column. "Штука" is used whenever the product has
    it among measure_units_names; falls back to "Блок" then "Коробка" only for the
    (checked-in-practice) minority of products that have no "Штука" unit at all.
    All available per-unit values are additionally preserved in raw.jsonl under
    "_all_units_weight" / "_all_units_size" (dict of unit_name -> value) so a future
    re-priority never requires a re-scrape.

Run: python3 scripts/scrape-sharik-figurnye-bukety.py
Re-run any time — it resumes from checkpoints and skips already-scraped products.
"""

import os
import re
import json
import time
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
import openpyxl
from openpyxl.utils import get_column_letter

BASE_URL = "https://new.sharik.ru"
CATEGORY_PATH = "/tovary-dly-prazdnika-cat/folgirovannye-shary/shary-figurnye-bukety/"
DETAIL_PATH = "/tovary-dly-prazdnika/{slug}/"

OUT_DIR = os.path.join(os.path.dirname(__file__), "..", "scraped-sharik-figurnye-bukety")
LISTING_FILE = os.path.join(OUT_DIR, "listing.json")
RAW_FILE = os.path.join(OUT_DIR, "raw.jsonl")
ERRORS_FILE = os.path.join(OUT_DIR, "errors.jsonl")
XLSX_OUT = os.path.join(
    os.path.dirname(__file__), "..", "All the Files with material here", "Sharik-figurnye-bukety.xlsx"
)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml",
    "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
}

INITIAL_STATE_RE = re.compile(r"window\.__INITIAL_STATE__\s*=\s*")

# Priority for picking the packaging row from the "Вид упаковки" table.
# See docstring above for rationale. Штука (per-piece) first — that's what's wanted.
PACKAGING_UNIT_PRIORITY = ["Штука", "Блок", "Коробка"]

COLUMNS = [
    "Название",
    "Артикул",
    "Ссылка",
    "Размеры/Габариты",
    "Цвет фольга",
    "Коллекция",
    "Торговая марка",
    "Вес брутто",
    "Размер",
]


def fetch_initial_state(url: str) -> dict:
    resp = requests.get(url, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    html = resp.text
    for s in re.findall(r"<script[^>]*>(.*?)</script>", html, re.S):
        if "__INITIAL_STATE__" not in s:
            continue
        m = INITIAL_STATE_RE.search(s)
        if not m:
            continue
        raw = s[m.end():].strip()
        decoder = json.JSONDecoder()
        obj, _ = decoder.raw_decode(raw)
        return obj
    raise ValueError(f"__INITIAL_STATE__ not found on {url}")


def fetch_with_retry(url: str, tries: int = 4, delay: float = 1.5) -> dict:
    last_err = None
    for i in range(tries):
        try:
            return fetch_initial_state(url)
        except Exception as e:
            last_err = e
            time.sleep(delay * (i + 1))
    raise last_err


# ---------- Phase 1: crawl listing ----------

def crawl_listing():
    if os.path.exists(LISTING_FILE):
        with open(LISTING_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        print(f"Listing checkpoint found: {len(data)} products already collected.")
        return data

    print("Fetching page 1 to determine total count...")
    obj = fetch_with_retry(f"{BASE_URL}{CATEGORY_PATH}?page=1")
    prod_data = obj["product"]["products"]
    count = prod_data["count"]
    page_size = len(prod_data["items"])
    total_pages = (count + page_size - 1) // page_size
    print(f"Total products: {count} across {total_pages} pages (page_size={page_size})")

    seen = {}
    for item in prod_data["items"]:
        seen[item["id"]] = {"id": item["id"], "slug": item["slug"], "code": item.get("code", "")}

    def fetch_page(page):
        obj = fetch_with_retry(f"{BASE_URL}{CATEGORY_PATH}?page={page}")
        return page, obj["product"]["products"]["items"]

    errors = []
    with ThreadPoolExecutor(max_workers=8) as pool:
        futures = {pool.submit(fetch_page, p): p for p in range(2, total_pages + 1)}
        done = 0
        for fut in as_completed(futures):
            page = futures[fut]
            done += 1
            try:
                _, items = fut.result()
                for item in items:
                    seen[item["id"]] = {"id": item["id"], "slug": item["slug"], "code": item.get("code", "")}
            except Exception as e:
                print(f"  ERROR page {page}: {e}")
                errors.append(page)
            if done % 20 == 0 or done == total_pages - 1:
                print(f"  {done}/{total_pages - 1} listing pages fetched -> {len(seen)} products so far")

    if errors:
        print(f"WARNING: {len(errors)} listing pages failed: {sorted(errors)}. Re-run to retry (checkpoint not yet saved).")
        raise SystemExit(1)

    data = list(seen.values())
    os.makedirs(OUT_DIR, exist_ok=True)
    with open(LISTING_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"Saved listing checkpoint: {len(data)} products -> {LISTING_FILE}")
    return data


# ---------- Phase 2: scrape product details ----------

def extract_raw(product: dict, url: str) -> dict:
    props = {p["name"]: p["value"] for p in product.get("properties", [])}
    orig = {p["name"]: p["value"] for p in product.get("origin_properties", [])}
    unit_names = product.get("measure_units_names") or []
    mup = {name: values for name, values in product.get("measure_units_properties", [])}

    def unit_value(prop_name, unit_name):
        values = mup.get(prop_name)
        if not values or unit_name not in unit_names:
            return ""
        idx = unit_names.index(unit_name)
        return values[idx] if idx < len(values) else ""

    chosen_unit = next((u for u in PACKAGING_UNIT_PRIORITY if u in unit_names), None)

    all_units_weight = {u: unit_value("Вес брутто", u) for u in unit_names}
    all_units_size = {u: unit_value("Размер", u) for u in unit_names}

    return {
        "id": product["id"],
        "Название": product.get("name", ""),
        "Артикул": product.get("code", ""),
        "Ссылка": url,
        "Размеры/Габариты": props.get("Размеры\\Габариты", ""),
        "Цвет фольга": props.get("Цвет фольга", ""),
        "Коллекция": props.get("Коллекция", ""),
        "Торговая марка": orig.get("Торговая марка", ""),
        "Вес брутто": unit_value("Вес брутто", chosen_unit),
        "Размер": unit_value("Размер", chosen_unit),
        "_chosen_unit": chosen_unit,
        "_unit_names": unit_names,
        "_all_units_weight": all_units_weight,
        "_all_units_size": all_units_size,
    }


def load_scraped_ids() -> set:
    ids = set()
    if os.path.exists(RAW_FILE):
        with open(RAW_FILE, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    ids.add(json.loads(line)["id"])
                except Exception:
                    pass
    return ids


def scrape_one(item: dict):
    url = f"{BASE_URL}{DETAIL_PATH.format(slug=item['slug'])}"
    obj = fetch_with_retry(url)
    product = obj["product"]["product"]
    if not product:
        raise ValueError("empty product state")
    return extract_raw(product, url)


def scrape_details(listing: list):
    scraped_ids = load_scraped_ids()
    todo = [it for it in listing if it["id"] not in scraped_ids]
    print(f"{len(scraped_ids)} already scraped, {len(todo)} remaining.")
    if not todo:
        return

    os.makedirs(OUT_DIR, exist_ok=True)
    written = 0
    failed = 0
    with open(RAW_FILE, "a", encoding="utf-8") as out_f, open(ERRORS_FILE, "a", encoding="utf-8") as err_f:
        with ThreadPoolExecutor(max_workers=8) as pool:
            futures = {pool.submit(scrape_one, it): it for it in todo}
            done_count = 0
            for fut in as_completed(futures):
                it = futures[fut]
                done_count += 1
                try:
                    row = fut.result()
                    out_f.write(json.dumps(row, ensure_ascii=False) + "\n")
                    out_f.flush()
                    written += 1
                except Exception as e:
                    print(f"  ERROR slug={it['slug']}: {e}")
                    err_f.write(json.dumps({"id": it["id"], "slug": it["slug"], "error": str(e)}, ensure_ascii=False) + "\n")
                    err_f.flush()
                    failed += 1
                if done_count % 50 == 0:
                    print(f"  {done_count}/{len(todo)} detail pages processed ({written} ok, {failed} failed)")
    print(f"Done. Wrote {written} new rows to {RAW_FILE} ({failed} failed, see {ERRORS_FILE})")


# ---------- Phase 3: build xlsx ----------

def build_xlsx():
    raws = []
    seen_ids = set()
    with open(RAW_FILE, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            row = json.loads(line)
            if row["id"] in seen_ids:
                continue
            seen_ids.add(row["id"])
            raws.append(row)

    raws.sort(key=lambda r: r["id"])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Шары фигурные букеты"

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    missing_report = []
    for i, row in enumerate(raws, start=2):
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=i, column=col_idx, value=row.get(col_name, ""))
            if col_name == "Артикул":
                cell.number_format = "@"
        missing = [c for c in COLUMNS if not row.get(c)]
        if missing:
            missing_report.append({"id": row["id"], "url": row.get("Ссылка", ""), "missing": missing})

    # reasonable column widths
    widths = [40, 14, 55, 20, 22, 20, 18, 20, 24]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    os.makedirs(os.path.dirname(XLSX_OUT), exist_ok=True)
    wb.save(XLSX_OUT)
    print(f"Wrote {len(raws)} rows -> {XLSX_OUT}")

    if missing_report:
        print(f"\n{len(missing_report)} products have at least one missing field:")
        for m in missing_report:
            print(f"  id={m['id']} missing={m['missing']} url={m['url']}")
    else:
        print("No missing fields across all products.")

    return raws, missing_report


def main():
    listing = crawl_listing()
    scrape_details(listing)
    raws, missing = build_xlsx()
    print(f"\nSummary: listing={len(listing)} listed, raw.jsonl={len(raws)} unique scraped rows.")


if __name__ == "__main__":
    main()
