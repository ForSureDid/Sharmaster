#!/usr/bin/env python3
"""
Rebuilds the final xlsx for the sharik.ru "Шары фигурные (букеты)" category from
the already-scraped checkpoints, adding the "Событие" column that the original
scrape-sharik-figurnye-bukety.py did not capture.

Inputs (both already on disk, no network calls in this script):
  - scraped-sharik-figurnye-bukety/raw.jsonl   (id, Название, Артикул, Ссылка,
    Размеры/Габариты, Цвет фольга, Коллекция, Торговая марка, Вес брутто, Размер)
  - scraped-sharik-figurnye-bukety/event.jsonl (id, Событие) — produced by
    scripts/scrape-sharik-figurnye-bukety-event.py

Output: All the Files with material here/Sharik-figurnye-bukety.xlsx
Column order: Название, Артикул, Ссылка, Размеры/Габариты, Цвет фольга,
              Коллекция, Событие, Торговая марка, Вес брутто, Размер
("Событие" placed right after "Коллекция" — both are thematic/grouping props).

Run: python3 scripts/build-sharik-figurnye-bukety-xlsx.py
"""

import os
import json

import openpyxl
from openpyxl.utils import get_column_letter

OUT_DIR = os.path.join(os.path.dirname(__file__), "..", "scraped-sharik-figurnye-bukety")
RAW_FILE = os.path.join(OUT_DIR, "raw.jsonl")
EVENT_FILE = os.path.join(OUT_DIR, "event.jsonl")
XLSX_OUT = os.path.join(
    os.path.dirname(__file__), "..", "All the Files with material here", "Sharik-figurnye-bukety.xlsx"
)

COLUMNS = [
    "Название",
    "Артикул",
    "Ссылка",
    "Размеры/Габариты",
    "Цвет фольга",
    "Коллекция",
    "Событие",
    "Торговая марка",
    "Вес брутто",
    "Размер",
]

WIDTHS = [40, 14, 55, 20, 22, 20, 18, 18, 20, 24]


def load_raw():
    raws = []
    seen_ids = set()
    with open(RAW_FILE, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            row = json.loads(line)
            if row["id"] in seen_ids:
                continue
            seen_ids.add(row["id"])
            raws.append(row)
    return raws


def load_events():
    events = {}
    if os.path.exists(EVENT_FILE):
        with open(EVENT_FILE, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                row = json.loads(line)
                events[row["id"]] = row.get("Событие", "")
    return events


def main():
    raws = load_raw()
    events = load_events()
    raws.sort(key=lambda r: r["id"])

    missing_event_ids = [r["id"] for r in raws if r["id"] not in events]
    if missing_event_ids:
        print(f"WARNING: {len(missing_event_ids)} products in raw.jsonl have no entry in event.jsonl yet "
              f"(run scripts/scrape-sharik-figurnye-bukety-event.py first). First few ids: {missing_event_ids[:10]}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Шары фигурные букеты"

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    filled_event = 0
    empty_event = 0
    for i, row in enumerate(raws, start=2):
        row = dict(row)
        row["Событие"] = events.get(row["id"], "")
        if row["Событие"]:
            filled_event += 1
        else:
            empty_event += 1
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=i, column=col_idx, value=row.get(col_name, ""))
            if col_name == "Артикул":
                cell.number_format = "@"

    for idx, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    os.makedirs(os.path.dirname(XLSX_OUT), exist_ok=True)
    wb.save(XLSX_OUT)

    print(f"Wrote {len(raws)} rows -> {XLSX_OUT}")
    print(f"Событие filled: {filled_event}")
    print(f"Событие empty: {empty_event}")


if __name__ == "__main__":
    main()
